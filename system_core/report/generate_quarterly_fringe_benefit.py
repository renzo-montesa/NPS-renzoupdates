import sys
from api.v1.employee_info.model import EmployeeInfoModel
from api.v1.payslip_entry_detail.model import PayslipEntryDetailModel
from api.v1.pay_element_info.model import PayElementInfoModel
from system_core.helper.array_to_dict import (
    array_to_dict_by_key,
    array_to_dict_by_keys_lower
)
from system_core.helper.create_full_name import create_full_name
from openpyxl import Workbook
from system_core.helper.report.format_cell import (
    format_range_font,
    format_range
)
from system_core.payslip_entry.compute_per_group import compute_per_group_by_key
from system_core.helper.join_fields import join_fields_with_key
from system_core.helper.get_distinct_keys import get_distinct_keys
from system_core.file_writer.write_to_file import write_json_to_file


REPORT_PATH = "C:\\Kirby\\Python\\Projects\\new_payroll_v2\\reports\\"


def generate_quarterly_fringe_benefit(data):
    print("Start | Generate quarterly report of fringe benefit tax", file=sys.stderr)

    """ Initialize options values """
    data = {}
    month = 1
    year = 2020

    """ Get company details """
    data['company'] = {
        'gross_up': 0.68,
        'fbt': 0.32
    }

    """ Get employee details """
    data['employees'] = array_to_dict_by_key(EmployeeInfoModel.get_employees_basic_info(), "employee_number")

    """ Get pay element info """
    data['pay_elements'] = array_to_dict_by_key(PayElementInfoModel.get_pay_elements(), "pay_element_code")

    """ Get list of fbt employee + earnings for the quarter """
    data['fbt_list'] = PayslipEntryDetailModel.get_fbt_employees_by_quarter(month, year)

    """ Get all fbt earnings """
    data['earnings'] = {}
    data['earnings']['first'] = array_to_dict_by_keys_lower(PayslipEntryDetailModel.get_fbt_earnings_by_month(month, year), ['employee_number', 'pay_element_code'])
    data['earnings']['second'] = array_to_dict_by_keys_lower(PayslipEntryDetailModel.get_fbt_earnings_by_month(month+1, year), ['employee_number', 'pay_element_code'])
    data['earnings']['third'] = array_to_dict_by_keys_lower(PayslipEntryDetailModel.get_fbt_earnings_by_month(month+2, year), ['employee_number', 'pay_element_code'])

    """ START | Create method to compute totals per grouping """
    """ Join employee info """
    data['earnings']['first'] = join_fields_with_key(data['earnings']['first'], data['employees'], 'employee_number', ['section_code', 'department_code'])
    data['earnings']['second'] = join_fields_with_key(data['earnings']['second'], data['employees'], 'employee_number', ['section_code', 'department_code'])
    data['earnings']['third'] = join_fields_with_key(data['earnings']['third'], data['employees'], 'employee_number', ['section_code', 'department_code'])

    """ Compute totals per department_code """
    data['earns_dept'] = {}
    data['earns_dept']['first'] = compute_per_group_by_key(data['earnings']['first'], 'department_code', 'amount')
    data['earns_dept']['second'] = compute_per_group_by_key(data['earnings']['second'], 'department_code', 'amount')
    data['earns_dept']['third'] = compute_per_group_by_key(data['earnings']['third'], 'department_code', 'amount')
    data['earns_dept']['list'] = get_distinct_keys([data['earns_dept']['first'], data['earns_dept']['second'], data['earns_dept']['third']])
    data['earns_dept']['list'].sort()

    """ Compute totals per pay_element_code """
    data['earns_code'] = {}
    data['earns_code']['first'] = compute_per_group_by_key(data['earnings']['first'], 'pay_element_code', 'amount')
    data['earns_code']['second'] = compute_per_group_by_key(data['earnings']['second'], 'pay_element_code', 'amount')
    data['earns_code']['third'] = compute_per_group_by_key(data['earnings']['third'], 'pay_element_code', 'amount')
    data['earns_code']['list'] = get_distinct_keys([data['earns_code']['first'], data['earns_code']['second'], data['earns_code']['third']])
    data['earns_code']['list'].sort()
    """ END | Create method to compute totals per grouping """

    """ Create excel file """
    """ Create workbook """
    wb = Workbook()

    """ Create sheet """
    ws = wb.active
    ws.title = "Fringe Benefit Report"

    """ Add header row """
    ws['A1'] = "Employee No."
    ws['B1'] = "Name"
    ws['C1'] = "Department"
    ws['D1'] = "Benefit"
    ws['E1'] = "January"
    ws['F1'] = "February"
    ws['G1'] = "March"
    ws['H1'] = "Total"
    ws['I1'] = "Gross Up Amount"
    ws['J1'] = "FBT"
    format_range_font(ws, 1, 1, 1, 10, bold=True)

    """ Format columns """
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 13
    #format_range(ws, 2, 3, row-1, 5, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")

    """ Add detail rows """
    row = 2
    for fbt in data['fbt_list']:
        fbt_index = fbt['employee_number'].lower() + fbt['pay_element_code'].lower()
        ws.cell(row=row, column=1, value=fbt['employee_number'])

        basic_info = data['employees'][fbt['employee_number']]
        ws.cell(row=row, column=2, value=create_full_name(basic_info['lastname'], basic_info['firstname'], basic_info['middlename']))

        ws.cell(row=row, column=3, value=basic_info['department_code'])

        if fbt['pay_element_code'] in data['pay_elements']:
            ws.cell(row=row, column=4, value=data['pay_elements'][fbt['pay_element_code']]['pay_element_description'])
        else:
            ws.cell(row=row, column=4, value=fbt['pay_element_code'])

        if fbt_index in data['earnings']['first']:
            ws.cell(row=row, column=5, value=data['earnings']['first'][fbt_index]['amount'])
        else:
            ws.cell(row=row, column=5, value=0)

        if fbt_index in data['earnings']['second']:
            ws.cell(row=row, column=6, value=data['earnings']['second'][fbt_index]['amount'])
        else:
            ws.cell(row=row, column=6, value=0)

        if fbt_index in data['earnings']['third']:
            ws.cell(row=row, column=7, value=data['earnings']['third'][fbt_index]['amount'])
        else:
            ws.cell(row=row, column=7, value=0)

        ws.cell(row=row, column=8).value = "=SUM(" + ws.cell(row=row, column=5).coordinate + ":" + ws.cell(row=row, column=7).coordinate + ")"
        ws.cell(row=row, column=9).value = "=" + ws.cell(row=row, column=8).coordinate + "/" + str(data['company']['gross_up'])
        ws.cell(row=row, column=10).value = "=" + ws.cell(row=row, column=9).coordinate + "*" + str(data['company']['fbt'])

        row += 1

    """ Add totals row """
    ws.cell(row=row, column=3).value = "Totals >>>>"
    ws.cell(row=row, column=8).value = "=SUM(" + ws.cell(row=2, column=8).coordinate + ":" + ws.cell(row=(row-1), column=8).coordinate + ")"
    ws.cell(row=row, column=9).value = "=" + ws.cell(row=row, column=8).coordinate + "/" + str(data['company']['gross_up'])
    ws.cell(row=row, column=10).value = "=" + ws.cell(row=row, column=9).coordinate + "*" + str(data['company']['fbt'])

    row += 4

    """ Add dept totals """
    dept_row_start = row
    for dept in data['earns_dept']['list']:
        ws.cell(row=row, column=4).value = "Total for " + dept + " >>>>"

        if dept in data['earns_dept']['first']:
            ws.cell(row=row, column=5, value=data['earns_dept']['first'][dept]['amount'])
        else:
            ws.cell(row=row, column=5, value=0)

        if dept in data['earns_dept']['second']:
            ws.cell(row=row, column=6, value=data['earns_dept']['second'][dept]['amount'])
        else:
            ws.cell(row=row, column=6, value=0)

        if dept in data['earns_dept']['third']:
            ws.cell(row=row, column=7, value=data['earns_dept']['third'][dept]['amount'])
        else:
            ws.cell(row=row, column=7, value=0)

        ws.cell(row=row, column=8).value = "=SUM(" + ws.cell(row=row, column=5).coordinate + ":" + ws.cell(row=row, column=7).coordinate + ")"
        ws.cell(row=row, column=9).value = "=" + ws.cell(row=row, column=8).coordinate + "/" + str(data['company']['gross_up'])
        ws.cell(row=row, column=10).value = "=" + ws.cell(row=row, column=9).coordinate + "*" + str(data['company']['fbt'])

        row += 1

    """ Add totals row """
    ws.cell(row=row, column=3).value = "Totals >>>>"
    ws.cell(row=row, column=8).value = "=SUM(" + ws.cell(row=dept_row_start, column=8).coordinate + ":" + ws.cell(row=(row-1), column=8).coordinate + ")"
    ws.cell(row=row, column=9).value = "=" + ws.cell(row=row, column=8).coordinate + "/" + str(data['company']['gross_up'])
    ws.cell(row=row, column=10).value = "=" + ws.cell(row=row, column=9).coordinate + "*" + str(data['company']['fbt'])
    row += 2

    """ Add code totals """
    code_row_start = row
    for code in data['earns_code']['list']:
        ws.cell(row=row, column=4).value = "Total for " + code + " >>>>"

        if code in data['earns_code']['first']:
            ws.cell(row=row, column=5, value=data['earns_code']['first'][code]['amount'])
        else:
            ws.cell(row=row, column=5, value=0)

        if code in data['earns_code']['second']:
            ws.cell(row=row, column=6, value=data['earns_code']['second'][code]['amount'])
        else:
            ws.cell(row=row, column=6, value=0)

        if code in data['earns_code']['third']:
            ws.cell(row=row, column=7, value=data['earns_code']['third'][code]['amount'])
        else:
            ws.cell(row=row, column=7, value=0)

        ws.cell(row=row, column=8).value = "=SUM(" + ws.cell(row=row, column=5).coordinate + ":" + ws.cell(row=row, column=7).coordinate + ")"
        ws.cell(row=row, column=9).value = "=" + ws.cell(row=row, column=8).coordinate + "/" + str(data['company']['gross_up'])
        ws.cell(row=row, column=10).value = "=" + ws.cell(row=row, column=9).coordinate + "*" + str(data['company']['fbt'])

        row += 1

    """ Add totals row """
    ws.cell(row=row, column=3).value = "Totals >>>>"
    ws.cell(row=row, column=8).value = "=SUM(" + ws.cell(row=code_row_start, column=8).coordinate + ":" + ws.cell(row=(row-1), column=8).coordinate + ")"
    ws.cell(row=row, column=9).value = "=" + ws.cell(row=row, column=8).coordinate + "/" + str(data['company']['gross_up'])
    ws.cell(row=row, column=10).value = "=" + ws.cell(row=row, column=9).coordinate + "*" + str(data['company']['fbt'])

    """ Format details rows """
    format_range(ws, 2, 5, row, 10, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")

    """ Save document """
    wb.save(REPORT_PATH + "quarterly_report_for_fringe_benefit.xlsx")

    print("Finished | Generate quarterly report of fringe benefit tax", file=sys.stderr)