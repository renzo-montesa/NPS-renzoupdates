from api.v1.report_field.model import ReportFieldModel
from api.v1.payslip_info.model import PayslipInfoModel
from api.v1.employee_info.model import EmployeeInfoModel
from api.v1.payslip_entry_detail.model import PayslipEntryDetailModel
from api.v1.timekeeping_detail.model import TimekeepingDetailModel
from api.v1.tk_element_info.model import TkElementInfoModel
from system_core.helper.array_to_dict import array_to_dict_by_key
from system_core.file_writer.write_to_file import write_json_to_file
from system_core.pay_element.get_pay_elements import get_pay_elements
from config import REPORT_DIR
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from system_core.helper.report.format_cell import (
    format_range,
    format_range_font
)

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders


def generate_payroll_report(data):
    data['report_fields'] = ReportFieldModel.get_all_by_filter(data['client_db'], report_id=data['report_id'])

    data['payslips'] = array_to_dict_by_key(PayslipInfoModel.get_payslips_by_payroll_period_id(data['client_db'], data['payroll_period_id']), 'employee_number')
    data['employees'] = array_to_dict_by_key(EmployeeInfoModel.get_all_employees(data['client_db']), 'employee_number')
    
    payslip_entries = PayslipEntryDetailModel.get_all_by_filter(data['client_db'], payroll_period_id=data['payroll_period_id'])

    payslip_entries_dict = {}
    for payslip_entry in payslip_entries:
        pe_key = payslip_entry['employee_number']
        if pe_key not in payslip_entries_dict:
            payslip_entries_dict[pe_key] = {}

        payslip_entries_dict[pe_key][payslip_entry['pay_element_code']] = payslip_entry['amount']
    data['payslip_entries'] = payslip_entries_dict

    timekeepings = TimekeepingDetailModel.get_all_by_filter(data['client_db'], payroll_period_id=data['payroll_period_id'])

    timekeepings_dict = {}
    for timekeeping in timekeepings:
        tk_key = timekeeping['employee_number']
        if tk_key not in timekeepings_dict:
            timekeepings_dict[tk_key] = {}

        timekeepings_dict[tk_key][timekeeping['tk_element_code']] = timekeeping['hours']
    data['timekeepings'] = timekeepings_dict


    get_pay_elements(data)
    tk_elements = TkElementInfoModel.get_all(data['client_db'])

    # Get ot codes
    ot_codes = []
    for tk_element in tk_elements:
        if tk_element['tk_element_type_code'] == 'overtime':
            ot_codes.append(tk_element['tk_element_code'])

    setup_fields = []
    for report_field in data['report_fields']:
        setup_fields.append(report_field['field'])

    pe_keys = sorted(PayslipEntryDetailModel.get_distinct_pec_by_ppid(data['client_db'], data['payroll_period_id']))


    # Get list of taxable and nontaxable codes
    taxable_codes = {}
    nontaxable_codes = {}

    for pe_key in pe_keys:
        if pe_key not in setup_fields:
            if pe_key in data['pay_elements']:
                if data['pay_elements'][pe_key]['pay_element_type_code'] == 'earning':
                    if 'is_taxable' in data['pay_elements'][pe_key]:
                        if data['pay_elements'][pe_key]['is_taxable'] == 'TRUE':
                            taxable_codes[pe_key] = {
                                'pay_element_code': data['pay_elements'][pe_key]['pay_element_code'],
                                'pay_element_description': data['pay_elements'][pe_key]['pay_element_description']
                            }
                        else:
                            nontaxable_codes[pe_key] = {
                                'pay_element_code': data['pay_elements'][pe_key]['pay_element_code'],
                                'pay_element_description': data['pay_elements'][pe_key]['pay_element_description']
                            }
                    else:
                        nontaxable_codes[pe_key] = {
                            'pay_element_code': data['pay_elements'][pe_key]['pay_element_code'],
                            'pay_element_description': data['pay_elements'][pe_key]['pay_element_description']
                        }

    # Get list of deduction codes
    deduction_codes = {}

    for pe_key in pe_keys:
        if pe_key not in setup_fields:
            if pe_key in data['pay_elements']:
                if data['pay_elements'][pe_key]['pay_element_type_code'] == 'deduction':
                    deduction_codes[pe_key] = {
                        'pay_element_code': data['pay_elements'][pe_key]['pay_element_code'],
                        'pay_element_description': data['pay_elements'][pe_key]['pay_element_description']
                    }


    report_fields = data['report_fields']
    payslips = data['payslips']
    employees = data['employees']
    payslip_entries = data['payslip_entries']
    timekeepings = data['timekeepings']


    # Compute totals (ot, taxable, nontax, gross, deductions)
    for employee_number, payslip_detail in payslip_entries.items():
        if 'total_taxable' not in payslips[employee_number]:
            payslips[employee_number]['total_taxable'] = 0.00

        if 'total_nontaxable' not in payslips[employee_number]:
            payslips[employee_number]['total_nontaxable'] = 0.00

        if 'gross_pay' not in payslips[employee_number]:
            payslips[employee_number]['gross_pay'] = 0.00

        if 'total_deduction' not in payslips[employee_number]:
            payslips[employee_number]['total_deduction'] = 0.00

        for pe_code in payslip_detail:
            if pe_code in data['pay_elements']:
                if data['pay_elements'][pe_code]['pay_element_type_code'] == 'earning':
                    if 'is_taxable' in data['pay_elements'][pe_code]:
                        if data['pay_elements'][pe_code]['is_taxable'] == 'TRUE':
                            payslips[employee_number]['total_taxable'] += (payslip_detail[pe_code] * data['pay_elements'][pe_code]['multiplier'])
                        else:
                            payslips[employee_number]['total_nontaxable'] += (payslip_detail[pe_code] * data['pay_elements'][pe_code]['multiplier'])
                    else:
                        payslips[employee_number]['total_nontaxable'] += (payslip_detail[pe_code] * data['pay_elements'][pe_code]['multiplier'])
                    payslips[employee_number]['gross_pay'] += (payslip_detail[pe_code] * data['pay_elements'][pe_code]['multiplier'])
                if data['pay_elements'][pe_code]['pay_element_type_code'] == 'deduction':
                    payslips[employee_number]['total_deduction'] += (payslip_detail[pe_code] * data['pay_elements'][pe_code]['multiplier'])


    write_json_to_file(data)


    # Generate excel report
    # Create workbook
    wb = Workbook()

    # Create sheet
    ws = wb.active
    ws.title = "Payroll Report"

    row = 1

    col_ctr = 1
    for report_field in report_fields:
        if report_field['description'] == 'AUTOFILL' and report_field['field'] == 'TAXABLE':
            for taxable_code in taxable_codes:
                ws.cell(row=row, column=col_ctr).value = taxable_codes[taxable_code]['pay_element_description']
                #ws.cell(row=row, column=col_ctr).value = taxable_codes[taxable_code]['pay_element_code']
                col_ctr += 1
            continue

        if report_field['description'] == 'AUTOFILL' and report_field['field'] == 'NONTAXABLE':
            for nontaxable_code in nontaxable_codes:
                ws.cell(row=row, column=col_ctr).value = nontaxable_codes[nontaxable_code]['pay_element_description']
                #ws.cell(row=row, column=col_ctr).value = nontaxable_codes[nontaxable_code]['pay_element_code']
                col_ctr += 1
            continue

        if report_field['description'] == 'AUTOFILL' and report_field['field'] == 'DEDUCTION':
            for deduction_code in deduction_codes:
                ws.cell(row=row, column=col_ctr).value = deduction_codes[deduction_code]['pay_element_description']
                #ws.cell(row=row, column=col_ctr).value = deduction_codes[deduction_code]['pay_element_code']
                col_ctr += 1
            continue

        ws.cell(row=row, column=col_ctr).value = report_field['description']
        col_ctr += 1

    # Format header row
    format_range_font(ws, 1, 1, 1, col_ctr, bold=True)

    row += 1

    for employee_number, employee in payslips.items():
        employee_info = employees[employee_number]
        middle_initial = "" if not employee_info['middlename'] else (employee_info['middlename'][0] + ".")
        employee_info['name'] = employee_info['lastname'] + ', ' + employee_info['firstname'] + ' ' + middle_initial if not employee_info['suffix'] else employee_info['lastname'] + ' ' + employee_info['suffix'] + ', ' + employee_info['firstname'] + ' ' + middle_initial

        col_ctr = 1
        for report_field in report_fields:
            if report_field['description'] == 'AUTOFILL' and report_field['field'] == 'TAXABLE':
                key = eval(report_field['key'])
                record = eval(report_field['tablename'] + "['" + key + "']")
                for taxable_code in taxable_codes:
                    pe_code = taxable_codes[taxable_code]['pay_element_code']
                    if pe_code in record:
                        value = eval("record['" + pe_code + "']")
                        ws.cell(row=row, column=col_ctr).value = value
                    else:
                        ws.cell(row=row, column=col_ctr).value = 0.00
                    col_ctr += 1
                continue

            if report_field['description'] == 'AUTOFILL' and report_field['field'] == 'NONTAXABLE':
                key = eval(report_field['key'])
                record = eval(report_field['tablename'] + "['" + key + "']")
                for nontaxable_code in nontaxable_codes:
                    pe_code = nontaxable_codes[nontaxable_code]['pay_element_code']
                    if pe_code in record:
                        value = eval("record['" + pe_code + "']")
                        ws.cell(row=row, column=col_ctr).value = value
                    else:
                        ws.cell(row=row, column=col_ctr).value = 0.00
                    col_ctr += 1
                continue

            if report_field['description'] == 'AUTOFILL' and report_field['field'] == 'DEDUCTION':
                key = eval(report_field['key'])
                record = eval(report_field['tablename'] + "['" + key + "']")
                for deduction_code in deduction_codes:
                    pe_code = deduction_codes[deduction_code]['pay_element_code']
                    if pe_code in record:
                        value = eval("record['" + pe_code + "']")
                        ws.cell(row=row, column=col_ctr).value = value
                    else:
                        ws.cell(row=row, column=col_ctr).value = 0.00
                    col_ctr += 1
                continue

            if report_field['key']:
                key = eval(report_field['key'])
                record = eval(report_field['tablename'] + "['" + key + "']")
                if report_field['field'] in record:
                    value = eval("record['" + report_field['field'] + "']")
                    ws.cell(row=row, column=col_ctr).value = value
                else:
                    ws.cell(row=row, column=col_ctr).value = 0.00
            else:
                value = eval(report_field['tablename'] + "['" + report_field['field'] + "']")
                ws.cell(row=row, column=col_ctr).value = value
            col_ctr += 1
        row += 1

    format_range(ws, 2, 6, row, col_ctr, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")

    if row > 2:
        ws.cell(row=row, column=1).value = '=COUNTA(A{0}:A{1})'.format(2, row-1)
        format_range(ws, row, 1, row+1, col_ctr, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")
        format_range_font(ws, row+1, 1, row+1, col_ctr, bold=True)

        row += 1

        for n in range(6, col_ctr):
            ws.cell(row=row, column=n).value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(n), 2, row-2)

    # for col in ws.columns:
    #     max_length = 0
    #     column = col[0].column
    #     for cell in col:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2) * 1.2
    #     ws.column_dimensions[column].width = adjusted_width

    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        adjusted_width = (length + 2) * 1.1
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Save document
    filename = "payroll_report_" + datetime.now().strftime('%Y%m%d%H%M%S') + ".xlsx"
    wb.save(REPORT_DIR + filename)


    print('Preparing email')

    # Sends out the E-mail with the attachment
    msg = MIMEMultipart()
    msg['Subject'] = 'Payroll Report'
    msg['From'] = 'NewPayrollSystem Email Sender'
    msg['To'] = data['email']

    # Open text file and attach text to message
    msg.attach(MIMEText('Please see the attached file for your generated report.', 'html'))

    # Setup the and attach the attachment
    datafile = open(REPORT_DIR + filename, "rb")
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(datafile.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename=" + filename)
    msg.attach(part)
    datafile.close()

    # Send email to recipients
    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.ehlo()
    server.login('kclopez@cpi-outsourcing.com', 'P4n1gur0 L4')
    server.sendmail('kclopez@cpi-outsourcing.com', data['email'], msg.as_string())
    server.quit()

    print('Email sent!')