import config
import sys

from openpyxl import Workbook
from system_core.helper.report.format_cell import (
    format_range,
    format_cell,
    align_range,
    align_cell_horizontal,
    format_range_font,
    format_border
)


REPORT_PATH = config.REPORT_DIR


def generate_summary_totals_comparison_report(data):
    print("Start | Generate summary totals comparison report", file=sys.stderr)

    # Generate excel report
    # Create workbook
    wb = Workbook()

    # Create sheet
    ws = wb.active
    ws.title = "Sheet1"

    # Add header
    ws['A1'] = "Run Date >> Jan 1, 2021 17:41:13"
    ws['C1'] = "Page 1"

    ws['A3'] = "Summary Totals Comparison Report"
    ws['A4'] = "ACER - ACER PHILIPPINES"
    ws['A5'] = "Payroll Period: 01/01/2021"

    # Add summary
    ws['A7'] = "Number of Period"
    ws['B7'] = "1"
    ws['A8'] = "Total Earnings"
    ws['B8'] = 100.00
    ws['A9'] = "Total Deductions"
    ws['B9'] = 100.00
    ws['A10'] = "Total Net Pay"
    ws['B10'] = 0.00
    ws['A11'] = "Difference"
    ws['B11'] = 0.00

    # Add orphan employees header
    ws['A13'] = "Bank Code"
    ws['B13'] = "Head Count"
    ws['C13'] = "Total Net Pay"
    format_range_font(ws, 13, 1, 13, 3, bold=True)

    # Add orphan employees details
    ws['A14'] = "BDO"
    ws['B14'] = 500.00
    ws['C14'] = 1000.00

    # Add orphan codes header
    ws['B16'] = "Orphan Codes"
    ws['A17'] = "Employee No"
    ws['B17'] = "Field"
    ws['C17'] = "Value"
    format_range_font(ws, 16, 1, 17, 3, bold=True)

    # Add orphan codes details
    ws['A18'] = "1234"
    ws['B18'] = "EARNCODE"
    ws['C18'] = "ALLOW"

    # Format columns
    # Set column widths
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 15

    # Merge cells
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)

    # Set font to bold
    format_range_font(ws, 1, 1, 5, 3, bold=True)
    format_range_font(ws, 7, 1, 11, 1, bold=True)

    # Set number format
    format_range(ws, 14, 2, 14, 3, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")

    # Set align
    align_range(ws, 1, 3, 1, 3, 0, 'right')
    align_range(ws, 3, 1, 5, 6, 0, 'center')
    align_range(ws, 7, 2, 7, 2, 0, 'right')

    # Save document
    wb.save(REPORT_PATH + "summary_totals_comparison_report.xlsx")

    print("Finished | Generate summary totals comparison report", file=sys.stderr)
