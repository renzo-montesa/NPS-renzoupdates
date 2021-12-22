import config
import sys

from openpyxl import Workbook
from system_core.helper.report.format_cell import (
    format_range,
    format_cell,
    align_range,
    align_cell_horizontal,
    format_range_font,
    format_border
)


REPORT_PATH = config.REPORT_DIR


def generate_no_current_payroll_report(data):
    print("Start | Generate list of employees with no current payroll report", file=sys.stderr)

    # Generate excel report
    # Create workbook
    wb = Workbook()

    # Create sheet
    ws = wb.active
    ws.title = "Sheet1"

    # Add header
    ws['A1'] = "Run Date >> Jan 1, 2021 17:41:13"
    ws['C1'] = "Page 1"

    ws['A3'] = "List of Employees with no Current Payroll"
    ws['A4'] = "ACER - ACER PHILIPPINES"
    ws['A5'] = "Payroll Period: 01/01/2021"

    # Add header
    ws['A7'] = "Employee No"
    ws['B7'] = "Name"
    ws['C7'] = "Status"

    # Add details
    ws['A8'] = "1234"
    ws['B8'] = "CRUZ, JUAN P."
    ws['C8'] = "2"

    # Format columns
    # Set column widths
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 14

    # Merge cells
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)

    # Set font to bold
    format_range_font(ws, 1, 1, 5, 3, bold=True)
    format_range_font(ws, 7, 1, 7, 3, bold=True)

    # Set align
    align_range(ws, 1, 3, 1, 3, 0, 'right')
    align_range(ws, 3, 1, 7, 3, 0, 'center')

    # Save document
    wb.save(REPORT_PATH + "no_current_payroll_report.xlsx")

    print("Finished | Generate list of employees with no current payroll report", file=sys.stderr)
