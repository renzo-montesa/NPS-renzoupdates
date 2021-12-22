import sys
from api.v1.employee_info.model import EmployeeInfoModel
from api.v1.payslip_entry_detail.model import PayslipEntryDetailModel
from system_core.file_writer.write_to_file import write_json_to_file
from system_core.helper.array_to_dict import (array_to_dict_by_key)
from system_core.helper.create_full_name import create_full_name
from system_core.helper.report.format_cell import (
    format_range,
    format_cell,
    align_range,
    align_cell_horizontal,
    format_range_font,
    format_border
)
from openpyxl import Workbook
from openpyxl.styles import Border, Side


REPORT_PATH = "C:\\Kirby\\Python\\Projects\\new_payroll_v2\\reports\\"


def generate_quarterly_earning_report(data):
    print("Start | Generate quarterly earning report", file=sys.stderr)

    """ Initialize options values """
    data = {}
    month = 1
    year = 2020

    """ Get employee details """
    data['employees'] = array_to_dict_by_key(EmployeeInfoModel.get_employees_basic_info(), "employee_number")

    """ Get employee list with earnings for the quarter """
    data['employee_list'] = PayslipEntryDetailModel.get_employees_with_earnings_by_quarter(month, year)

    """ Get sum amount per month as first, second, third """
    data['earnings'] = {}
    data['earnings']['first'] = array_to_dict_by_key(PayslipEntryDetailModel.get_earnings_by_month(month, year), 'employee_number')
    data['earnings']['second'] = array_to_dict_by_key(PayslipEntryDetailModel.get_earnings_by_month(month+1, year), 'employee_number')
    data['earnings']['third'] = array_to_dict_by_key(PayslipEntryDetailModel.get_earnings_by_month(month+2, year), 'employee_number')

    """ Generate excel report """
    """ Create workbook """
    wb = Workbook()

    """ Create sheet """
    ws = wb.active
    ws.title = "1st Qtr " + str(year)

    """ Add header row """
    ws['A1'] = "Employee No"
    ws['B1'] = "Name"
    ws['C1'] = "January"
    ws['D1'] = "February"
    ws['E1'] = "March"
    align_range(ws, 1, 1, 1, 5, 0, 'center')

    """ Add details """
    row = 2
    for employee in data['employee_list']:
        ws.cell(row=row, column=1, value=employee['employee_number'])

        if employee['employee_number'] in data['employees']:
            basic_info = data['employees'][employee['employee_number']]
            ws.cell(row=row, column=2, value=create_full_name(basic_info['lastname'], basic_info['firstname'], basic_info['middlename']))

        if employee['employee_number'] in data['earnings']['first']:
            ws.cell(row=row, column=3, value=data['earnings']['first'][employee['employee_number']]['amount'])
        else:
            ws.cell(row=row, column=3, value=0)

        if employee['employee_number'] in data['earnings']['second']:
            ws.cell(row=row, column=4, value=data['earnings']['second'][employee['employee_number']]['amount'])
        else:
            ws.cell(row=row, column=4, value=0)

        if employee['employee_number'] in data['earnings']['third']:
            ws.cell(row=row, column=5, value=data['earnings']['third'][employee['employee_number']]['amount'])
        else:
            ws.cell(row=row, column=5, value=0)

        row += 1
    format_border(ws, ws.cell(row=1, column=1).coordinate + ":" + ws.cell(row=(row-1), column=5).coordinate)

    """ Format columns """
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    format_range(ws, 2, 3, row-1, 5, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")

    """ Add totals """
    row += 2
    ws.cell(row=row, column=1, value="Per Month Totals")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    align_cell_horizontal(ws, row, 1, 'center')
    ws.cell(row=row, column=3).value = "=SUM(" + ws.cell(row=2, column=3).coordinate + ":" + ws.cell(row=(row-3), column=3).coordinate + ")"
    ws.cell(row=row, column=4).value = "=SUM(" + ws.cell(row=2, column=4).coordinate + ":" + ws.cell(row=(row-3), column=4).coordinate + ")"
    ws.cell(row=row, column=5).value = "=SUM(" + ws.cell(row=2, column=5).coordinate + ":" + ws.cell(row=(row-3), column=5).coordinate + ")"
    format_range(ws, row, 3, row, 5, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")
    format_border(ws, ws.cell(row=row, column=1).coordinate + ":" + ws.cell(row=row, column=5).coordinate)

    """ Add grand total """
    row += 2
    ws.cell(row=row, column=2, value="GRAND TOTAL")
    formula = "=" + ws.cell(row=(row-2), column=3).coordinate + "+" + ws.cell(row=(row-2), column=4).coordinate + "+" + ws.cell(row=(row-2), column=5).coordinate
    ws.cell(row=row, column=3).value = formula
    format_cell(ws, row, 3, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")
    format_range_font(ws, row, 2, row, 3, bold=True)
    ws.cell(row=row, column=3).border = Border(bottom=Side(border_style='double', color='000000'))

    """ Add footer """
    row += 3
    ws.cell(row=row, column=1, value="Prepared by:")
    row += 5
    ws.cell(row=row, column=1, value="________________________________")

    """ Save document """
    wb.save(REPORT_PATH + "quarterly_earnings_report.xlsx")

    print("Finished | Generate quarterly earning report", file=sys.stderr)