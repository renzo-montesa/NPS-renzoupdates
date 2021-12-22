import config
import sys

from openpyxl import Workbook
from system_core.helper.report.format_cell import (
    format_range,
    format_cell,
    align_range,
    align_cell_horizontal,
    format_range_font,
    format_border
)


REPORT_PATH = config.REPORT_DIR


def generate_loans_exception_report(data):
    print("Start | Generate loans exception report", file=sys.stderr)

    # Generate excel report
    # Create workbook
    wb = Workbook()

    # Create sheet
    ws = wb.active
    ws.title = "Loans Exception Report"

    # Add header
    ws['A1'] = "Run Date >> Jan 1, 2021 17:41:13"
    ws['F1'] = "Page 1"

    ws['A3'] = "Loans Exception Report"
    ws['A4'] = "ACER - ACER PHILIPPINES"
    ws['A5'] = "Payroll Period: 01/01/2021"

    ws['A7'] = "Employee No"
    ws['B7'] = "Name"
    ws['C7'] = "Code"
    ws['D7'] = "Date of Loan"
    ws['E7'] = "Payment Amount"
    ws['F7'] = "Deduction Amount"

    # Add details
    ws['A8'] = "1234"
    ws['B8'] = "CRUZ, JOHN P."
    ws['C8'] = "SSSLN"
    ws['D8'] = "01/01/2020"
    ws['E8'] = 150.00
    ws['F8'] = 200.00

    # Format columns
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # Merge cells
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=6)

    # Set font to bold
    format_range_font(ws, 1, 1, 7, 6, bold=True)

    # Set number format
    format_range(ws, 8, 5, 8, 6, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")

    # Set align
    align_range(ws, 1, 6, 1, 6, 0, 'right')
    align_range(ws, 3, 1, 5, 6, 0, 'center')
    align_range(ws, 7, 1, 7, 6, 0, 'center')

    # Save document
    wb.save(REPORT_PATH + "loans_exception_report.xlsx")

    print("Finished | Generate loans exception report", file=sys.stderr)
