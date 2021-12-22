import config
import sys

from openpyxl import Workbook
from system_core.helper.report.format_cell import (
    format_range,
    format_cell,
    align_range,
    align_cell_horizontal,
    format_range_font,
    format_border
)


REPORT_PATH = config.REPORT_DIR


def generate_atu_ot_exempt_report(data):
    print("Start | Generate atu and ot exception report", file=sys.stderr)

    # Generate excel report
    # Create workbook
    wb = Workbook()

    # Create sheet
    ws = wb.active
    ws.title = "Sheet1"

    # Add header
    ws['A1'] = "Run Date >> Jan 1, 2021 17:41:13"
    ws['C1'] = "Page 1"

    ws['A3'] = "ATU & OT Exception Report"
    ws['A4'] = "ACER - ACER PHILIPPINES"
    ws['A5'] = "Payroll Period: 01/01/2021"

    # Add header
    ws['A7'] = "Employee No"
    ws['B7'] = "Name"
    ws['C7'] = "No of Hours"

    # Add group header
    ws['A8'] = "ABSENT"
    format_range_font(ws, 8, 1, 8, 3, bold=True)

    # Add group details
    ws['A9'] = "1234"
    ws['B9'] = "CRUZ, JUAN P."
    ws['C9'] = 8.00

    # Add group header
    ws['A11'] = "TARDINESS"
    format_range_font(ws, 11, 1, 11, 3, bold=True)

    # Add group details
    ws['A12'] = "1234"
    ws['B12'] = "CRUZ, JUAN P."
    ws['C12'] = 3.00

    # Add group header
    ws['A14'] = "UNDERTIME"
    format_range_font(ws, 14, 1, 14, 3, bold=True)

    # Add group details
    ws['A15'] = "1234"
    ws['B15'] = "CRUZ, JUAN P."
    ws['C15'] = 1.65

    # Add group header
    ws['A17'] = "REG_OT_HRS"
    format_range_font(ws, 17, 1, 17, 3, bold=True)

    # Add group details
    ws['A18'] = "1234"
    ws['B18'] = "CRUZ, JUAN P."
    ws['C18'] = 15.00

    # Format columns
    # Set column widths
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 14

    # Merge cells
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)

    # Set font to bold
    format_range_font(ws, 1, 1, 7, 3, bold=True)

    # Set align
    align_range(ws, 1, 3, 1, 3, 0, 'right')
    align_range(ws, 3, 1, 7, 3, 0, 'center')

    # Save document
    wb.save(REPORT_PATH + "atu_ot_exempt_report.xlsx")

    print("Finished | Generate atu and ot exception report", file=sys.stderr)
