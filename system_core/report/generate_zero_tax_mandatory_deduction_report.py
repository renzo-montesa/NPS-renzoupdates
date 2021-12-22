import config
import sys

from openpyxl import Workbook
from system_core.helper.report.format_cell import (
    format_range,
    format_cell,
    align_range,
    align_cell_horizontal,
    format_range_font,
    format_border
)


REPORT_PATH = config.REPORT_DIR


def generate_zero_tax_mandatory_deduction_report(data):
    print("Start | Generate zero tax and mandatory deduction report", file=sys.stderr)

    # Generate excel report
    # Create workbook
    wb = Workbook()

    # Create sheet
    ws = wb.active
    ws.title = "Sheet1"

    # Add header
    ws['A1'] = "Run Date >> Jan 1, 2021 17:41:13"
    ws['C1'] = "Page 1"

    ws['A3'] = "List of Employees with Zero Tax and Mandatory Deduction"
    ws['A4'] = "ACER - ACER PHILIPPINES"
    ws['A5'] = "Payroll Period: 01/01/2021"

    # Add header
    ws['A7'] = "Employee No"
    ws['B7'] = "Name"
    ws['C7'] = "Amount"
    format_range_font(ws, 7, 1, 7, 3, bold=True)

    # Add subgroup
    ws['A8'] = "W_TAX"
    format_range_font(ws, 8, 1, 8, 3, bold=True)
    ws['A9'] = "1234"
    ws['B9'] = "CRUZ, JUAN P."
    ws['C9'] = 100.00

    # Add subgroup
    ws['A11'] = "EMPL_SSS"
    format_range_font(ws, 11, 1, 11, 3, bold=True)
    ws['A12'] = "1234"
    ws['B12'] = "CRUZ, JUAN P."
    ws['C12'] = 100.00

    # Add subgroup
    ws['A14'] = "EMPR_SSS"
    format_range_font(ws, 14, 1, 14, 3, bold=True)
    ws['A15'] = "1234"
    ws['B15'] = "CRUZ, JUAN P."
    ws['C15'] = 100.00

    # Add subgroup
    ws['A17'] = "ECC"
    format_range_font(ws, 17, 1, 17, 3, bold=True)
    ws['A18'] = "1234"
    ws['B18'] = "CRUZ, JUAN P."
    ws['C18'] = 100.00

    # Add subgroup
    ws['A20'] = "EMPL_MED"
    format_range_font(ws, 20, 1, 20, 3, bold=True)
    ws['A21'] = "1234"
    ws['B21'] = "CRUZ, JUAN P."
    ws['C21'] = 100.00

    # Add subgroup
    ws['A23'] = "EMPR_MED"
    format_range_font(ws, 23, 1, 23, 3, bold=True)
    ws['A24'] = "1234"
    ws['B24'] = "CRUZ, JUAN P."
    ws['C24'] = 100.00

    # Add subgroup
    ws['A26'] = "EMPL_PAGIB"
    format_range_font(ws, 26, 1, 26, 3, bold=True)
    ws['A27'] = "1234"
    ws['B27'] = "CRUZ, JUAN P."
    ws['C27'] = 100.00

    # Add subgroup
    ws['A29'] = "EMPR_PAGIB"
    format_range_font(ws, 29, 1, 29, 3, bold=True)
    ws['A30'] = "1234"
    ws['B30'] = "CRUZ, JUAN P."
    ws['C30'] = 100.00

    # Format columns
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 20

    # Merge cells
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)

    # Set font to bold
    format_range_font(ws, 1, 1, 5, 3, bold=True)

    # Set number format
    format_range(ws, 8, 3, 30, 3, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")

    # Set align
    align_range(ws, 1, 3, 1, 3, 0, 'right')
    align_range(ws, 3, 1, 5, 6, 0, 'center')

    # Save document
    wb.save(REPORT_PATH + "zero_tax_mandatory_deduction_report.xlsx")

    print("Finished | Generate zero tax and mandatory deduction report", file=sys.stderr)
