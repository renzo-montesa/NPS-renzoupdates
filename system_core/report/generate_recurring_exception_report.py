import config
import sys

from openpyxl import Workbook
from system_core.helper.report.format_cell import (
    format_range,
    format_cell,
    align_range,
    align_cell_horizontal,
    format_range_font,
    format_border
)


REPORT_PATH = config.REPORT_DIR


def generate_recurring_exception_report(data):
    print("Start | Generate recurring exception report", file=sys.stderr)

    # Generate excel report
    # Create workbook
    wb = Workbook()

    # Create sheet
    ws = wb.active
    ws.title = "Sheet1"

    # Add header
    ws['A1'] = "Run Date >> Jan 1, 2021 17:41:13"
    ws['E1'] = "Page 1"

    ws['A3'] = "Recurring Earnings and Deductions Exception Report"
    ws['A4'] = "ACER - ACER PHILIPPINES"
    ws['A5'] = "Payroll Period: 01/01/2021"

    # Add group header (Earnings)
    ws['A7'] = "Employee No"
    ws['B7'] = "Name"
    ws['C7'] = "Code"
    ws['D7'] = "Recurring Amount"
    ws['E7'] = "One-time Amount"
    ws['A8'] = "Earnings"

    # Add details (Earnings)
    ws['A9'] = "1234"
    ws['B9'] = "CRUZ, JOHN P."
    ws['C9'] = "TRANSPO"
    ws['D9'] = 150.00
    ws['E9'] = 125.00

    # Add group header (Deductions)
    ws['A11'] = "Deductions"

    # Add details (Deductions)
    ws['A12'] = "1234"
    ws['B12'] = "CRUZ, JOHN P."
    ws['C12'] = "COMPDED"
    ws['D12'] = 150.00
    ws['E12'] = 125.00

    # Add summary totals
    ws['A15'] = "Total Earnings Count:"
    ws['B15'] = 1.00
    ws['A16'] = "Total Earnings Amount from Earndetl:"
    ws['B16'] = 125.00
    ws['A17'] = "Total Earnings Amount from Rec_earn:"
    ws['B17'] = 150.00
    ws['A18'] = "Total Orphaned Earnings Amount from Earndetl:"
    ws['B18'] = 100.00
    ws['A19'] = "Total Orphaned Earnings Count from Earndetl:"
    ws['B19'] = 1.00

    ws['A21'] = "Total Deductions Count:"
    ws['B21'] = 1.00
    ws['A22'] = "Total Deductions Amount from Ded_detl:"
    ws['B22'] = 125.00
    ws['A23'] = "Total Deductions Amount from Rec_ded:"
    ws['B23'] = 150.00
    ws['A24'] = "Total Orphaned Deductions Amount from Ded_detl:"
    ws['B24'] = 100.00
    ws['A25'] = "Total Orphaned Deductions Count from Ded_detl:"
    ws['B25'] = 1.00

    # Format columns
    # Set column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 31
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17

    # Merge cells
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)

    # Set font to bold
    format_range_font(ws, 1, 1, 7, 5, bold=True)
    format_range_font(ws, 8, 1, 8, 1, bold=True)
    format_range_font(ws, 11, 1, 11, 1, bold=True)

    # Set number format
    format_range(ws, 9, 4, 12, 5, "_(#,##0.00_);_((#,##0.00);_(-??_);_(@_)")

    # Set align
    align_range(ws, 1, 5, 1, 5, 0, 'right')
    align_range(ws, 3, 1, 5, 5, 0, 'center')
    align_range(ws, 7, 1, 7, 5, 0, 'center')

    # Save document
    wb.save(REPORT_PATH + "recurring_earnings_deductions_exception_report.xlsx")

    print("Finished | Generate recurring exception report", file=sys.stderr)
