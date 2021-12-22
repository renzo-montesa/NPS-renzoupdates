from openpyxl.styles import Alignment, Font, Border, Side


def format_range(sheet, x1, y1, x2, y2, format):
    i = x1
    while i <= x2:
        j = y1
        while j <= y2:
            format_cell(sheet, i, j, format)
            j += 1
        i += 1


def format_cell(sheet, x, y, format):
    sheet.cell(row=x, column=y).number_format = format


# Function: align_range
# Align given range
# direction: 0 - horizontal; 1 - vertical
def align_range(sheet, x1, y1, x2, y2, direction, alignment):
    i = x1
    while i <= x2:
        j = y1
        while j <= y2:
            if direction == 0:
                align_cell_horizontal(sheet, i, j, alignment)
            elif direction == 1:
                align_cell_vertical(sheet, i, j, alignment)
            j += 1
        i += 1


def align_cell_horizontal(sheet, row, column, alignment):
    sheet.cell(row=row, column=column).alignment = Alignment(horizontal=alignment)


def align_cell_vertical(sheet, row, column, alignment):
    sheet.cell(row=row, column=column).alignment = Alignment(vertical=alignment)


def format_range_font(sheet, x1, y1, x2, y2, **kwargs):
    i = x1
    while i <= x2:
        j = y1
        while j <= y2:
            format_cell_font(sheet, i, j, **kwargs)
            j += 1
        i += 1

    
def format_cell_font(sheet, x, y, **kwargs):
    sheet.cell(row=x, column=y).font = Font(**kwargs)


def format_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border